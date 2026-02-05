import io
import re
import zipfile
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import MergedCell
from copy import copy as pycopy


# =========================
# Utils
# =========================
def s(x) -> str:
    return str(x).strip()


def sl(x) -> str:
    return str(x).strip().lower()


def colname(c) -> str:
    if isinstance(c, tuple):
        return str(c[0])
    return str(c)


def normalize_int(val: Any) -> Optional[int]:
    if val is None:
        return None
    if isinstance(val, float) and pd.isna(val):
        return None
    if isinstance(val, (int, float)):
        try:
            return int(round(float(val)))
        except Exception:
            return None
    txt = str(val)
    digits = re.sub(r"[^\d]", "", txt)
    return int(digits) if digits.isdigit() else None


def dedupe_columns(cols: List[Any]) -> List[Optional[str]]:
    out: List[Optional[str]] = []
    seen: Dict[str, int] = {}
    for c in cols:
        if c is None or (isinstance(c, float) and pd.isna(c)):
            out.append(None)
            continue
        name = s(c)
        if name not in seen:
            seen[name] = 0
            out.append(name)
        else:
            seen[name] += 1
            out.append(f"{name}_{seen[name]}")
    return out


def find_col_contains(df: pd.DataFrame, patterns: List[str]) -> Optional[str]:
    cols = [colname(c) for c in df.columns]
    cols_l = [sl(c) for c in cols]
    for p in patterns:
        p_l = sl(p)
        for i, c_l in enumerate(cols_l):
            if p_l in c_l:
                return cols[i]
    return None


# =========================
# SKU parser
# =========================
def parse_platform_sku(full_sku: Any) -> Tuple[str, List[str]]:
    if full_sku is None or (isinstance(full_sku, float) and pd.isna(full_sku)):
        return "", []
    parts = s(full_sku).split("+")
    base = parts[0].strip()
    addons = [p.strip() for p in parts[1:] if p.strip()]
    return base, addons


# =========================
# Read Pricelist smart header
# =========================
def read_pricelist_smart(file, sheet_name=0) -> pd.DataFrame:
    raw = pd.read_excel(file, sheet_name=sheet_name, header=None)

    header_idx = None
    for i in range(min(25, len(raw))):
        row = [sl(x) for x in raw.iloc[i].tolist()]
        if "kodebarang" in row or "kode barang" in row:
            header_idx = i
            break

    if header_idx is None:
        st.error("❌ Header Pricelist tidak ditemukan (kolom KODEBARANG).")
        st.stop()

    df = raw.iloc[header_idx:].copy()
    df.columns = dedupe_columns([colname(c) for c in df.iloc[0].tolist()])
    df = df.iloc[1:].reset_index(drop=True)
    df = df.loc[:, df.columns.notna()]
    return df


# =========================
# Scale auto-detect (AUTO x1000 atau tidak)
# =========================
def detect_price_scale_from_pricelist(df_pl: pd.DataFrame, price_col: str) -> int:
    nums = []
    for v in df_pl[price_col].head(200).tolist():
        iv = normalize_int(v)
        if iv is not None and iv > 0:
            nums.append(iv)

    if not nums:
        return 1000

    nums.sort()
    median = nums[len(nums) // 2]
    return 1000 if median < 100000 else 1


# =========================
# Build maps
# =========================
@dataclass
class Rules:
    price_scale: int
    discount_rp: int


def build_base_maps(
    df_pl: pd.DataFrame,
    col_sku: str,
    col_price: str,
    col_stock_tot: str,
    price_scale: int
) -> Tuple[Dict[str, int], Dict[str, int]]:
    price_map: Dict[str, int] = {}
    stock_map: Dict[str, int] = {}

    for _, r in df_pl.iterrows():
        sku = s(r.get(col_sku, ""))
        if not sku:
            continue

        price_raw = normalize_int(r.get(col_price, None))
        if price_raw is not None:
            price_map[sku] = int(price_raw) * int(price_scale)

        stock_raw = normalize_int(r.get(col_stock_tot, None))
        if stock_raw is not None:
            stock_map[sku] = int(stock_raw)

    return price_map, stock_map


def build_addon_map(df_add: pd.DataFrame, price_scale: int) -> Dict[str, int]:
    code_col = (
        find_col_contains(df_add, ["standarisasi kode sku di varian"]) or
        find_col_contains(df_add, ["addon_code"]) or
        find_col_contains(df_add, ["kode sku"]) or
        find_col_contains(df_add, ["kode"])
    )
    price_col = find_col_contains(df_add, ["harga"]) or find_col_contains(df_add, ["price"])

    if not code_col or not price_col:
        st.error("❌ Kolom kode/harga tidak ditemukan di Addon Mapping.")
        st.stop()

    m: Dict[str, int] = {}
    for _, r in df_add.iterrows():
        code = s(r.get(code_col, "")).upper()
        if not code:
            continue
        val = normalize_int(r.get(price_col, None))
        if val is None:
            continue
        m[code] = int(val) * int(price_scale)

    return m


def calc_addon_total(addons: List[str], addon_map: Dict[str, int]) -> Tuple[int, List[str]]:
    total = 0
    missing = []
    for a in addons:
        key = s(a).upper()
        if key in addon_map:
            total += addon_map[key]
        else:
            missing.append(s(a))
    return total, missing


# =========================
# openpyxl helpers
# =========================
def merged_topleft(ws, row: int, col: int) -> Tuple[int, int]:
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng.min_row, rng.min_col
    return row, col


def set_cell_value_safe(ws, row: int, col: int, value: Any):
    r0, c0 = merged_topleft(ws, row, col)
    cell = ws.cell(row=r0, column=c0)
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def copy_cell_safe(src_ws, dst_ws, r: int, c: int):
    src_cell = src_ws.cell(row=r, column=c)
    dst_cell = dst_ws.cell(row=r, column=c)

    if isinstance(src_cell, MergedCell) or isinstance(dst_cell, MergedCell):
        return

    dst_cell.value = src_cell.value
    if src_cell.has_style:
        dst_cell._style = pycopy(src_cell._style)
    dst_cell.number_format = src_cell.number_format
    dst_cell.font = pycopy(src_cell.font)
    dst_cell.fill = pycopy(src_cell.fill)
    dst_cell.border = pycopy(src_cell.border)
    dst_cell.alignment = pycopy(src_cell.alignment)
    dst_cell.protection = pycopy(src_cell.protection)
    dst_cell.comment = src_cell.comment


def copy_sheet_structure(src_ws, dst_ws):
    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width
        dst_ws.column_dimensions[col_letter].hidden = dim.hidden
        dst_ws.column_dimensions[col_letter].outlineLevel = dim.outlineLevel

    for r, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[r].height = dim.height
        dst_ws.row_dimensions[r].hidden = dim.hidden
        dst_ws.row_dimensions[r].outlineLevel = dim.outlineLevel

    dst_ws.freeze_panes = src_ws.freeze_panes

    try:
        for merged in list(src_ws.merged_cells.ranges):
            dst_ws.merge_cells(str(merged))
    except Exception:
        pass


def find_header_row_and_cols(ws, scan_rows: int = 30) -> Tuple[int, Dict[str, int]]:
    for r in range(1, scan_rows + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        vals_l = [sl(v) for v in vals]

        hits = {}
        if "sku" in vals_l:
            hits["sku"] = vals_l.index("sku") + 1
        if "harga" in vals_l:
            hits["harga"] = vals_l.index("harga") + 1
        if "stok" in vals_l:
            hits["stok"] = vals_l.index("stok") + 1

        if "sku" in hits and "harga" in hits:
            return r, hits

    return -1, {}


def header_signature(ws, header_row: int) -> List[str]:
    vals = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
    return [sl(v) for v in vals]


# =========================
# Marketplace detect
# =========================
def detect_marketplace_from_filename(filename: str) -> Optional[str]:
    name = sl(filename)
    if "shopee" in name:
        return "shopee"
    if "tiktok" in name:
        return "tiktok"
    return None


# =========================
# Filtering eligible rows
# =========================
def collect_rows_to_update(
    src_ws,
    header_row: int,
    sku_col: int,
    base_price_map: Dict[str, int],
    addon_map: Dict[str, int],
    issues_file: str
) -> Tuple[List[int], List[Dict[str, Any]]]:
    selected_rows = []
    issues = []

    empty_run = 0
    for r in range(header_row + 1, src_ws.max_row + 1):
        sku_val = src_ws.cell(row=r, column=sku_col).value

        if sku_val is None or s(sku_val) == "":
            empty_run += 1
            if empty_run >= 50:
                break
            continue
        empty_run = 0

        base, addons = parse_platform_sku(sku_val)
        base_price = base_price_map.get(base)

        if base_price is None:
            # SKU base tidak ketemu -> tidak masuk hasil, tidak error
            continue

        addon_total, missing = calc_addon_total(addons, addon_map)
        if missing:
            # aturan kamu: ada 1 addon tidak ketemu -> jangan ubah apapun
            issues.append({
                "file": issues_file,
                "row": r,
                "sku_full": s(sku_val),
                "base_sku": base,
                "reason": f"SKIP (addon tidak ketemu): {','.join(missing)}"
            })
            continue

        selected_rows.append(r)

    return selected_rows, issues


def append_row_copy(src_ws, dst_ws, src_row: int, dst_row: int, max_col: int):
    dst_ws.row_dimensions[dst_row].height = src_ws.row_dimensions[src_row].height
    for c in range(1, max_col + 1):
        src_cell = src_ws.cell(row=src_row, column=c)
        dst_cell = dst_ws.cell(row=dst_row, column=c)

        if isinstance(src_cell, MergedCell) or isinstance(dst_cell, MergedCell):
            continue

        dst_cell.value = src_cell.value
        if src_cell.has_style:
            dst_cell._style = pycopy(src_cell._style)
        dst_cell.number_format = src_cell.number_format
        dst_cell.font = pycopy(src_cell.font)
        dst_cell.fill = pycopy(src_cell.fill)
        dst_cell.border = pycopy(src_cell.border)
        dst_cell.alignment = pycopy(src_cell.alignment)
        dst_cell.protection = pycopy(src_cell.protection)
        dst_cell.comment = src_cell.comment


def create_template_from_first_file(first_bytes: bytes) -> Tuple[Workbook, Any, int, Dict[str, int], List[str]]:
    wb = load_workbook(io.BytesIO(first_bytes))
    ws = wb.worksheets[0]
    header_row, cols = find_header_row_and_cols(ws, scan_rows=30)
    if header_row == -1:
        raise ValueError("Header Mass Update tidak ketemu (butuh kolom SKU & Harga).")

    sig = header_signature(ws, header_row)
    return wb, ws, header_row, cols, sig


def build_combined_workbook_from_template(template_ws, header_row: int) -> Tuple[Workbook, Any]:
    dst_wb = Workbook()
    default_ws = dst_wb.active
    dst_wb.remove(default_ws)

    dst_ws = dst_wb.create_sheet(title=template_ws.title)

    copy_sheet_structure(template_ws, dst_ws)

    max_col = template_ws.max_column
    for r in range(1, header_row + 1):
        dst_ws.row_dimensions[r].height = template_ws.row_dimensions[r].height
        for c in range(1, max_col + 1):
            copy_cell_safe(template_ws, dst_ws, r, c)

    return dst_wb, dst_ws


# =========================
# UI
# =========================
st.set_page_config(page_title="Web App Update Harga", layout="wide")
st.title("Web App Update Harga")

c1, c2, c3 = st.columns(3)
with c1:
    mass_files = st.file_uploader("Upload Mass Update (bisa banyak)", type=["xlsx"], accept_multiple_files=True)
with c2:
    pl_file = st.file_uploader("Upload Pricelist", type=["xlsx"])
with c3:
    addon_file = st.file_uploader("Upload Addon Mapping", type=["xlsx"])

st.divider()

left, right = st.columns([2, 1])
with left:
    output_mode = st.radio(
        "Output",
        options=["Gabung jadi 1 file (untuk upload omnichannel)", "Tetap per file (ZIP)"],
        index=0
    )
with right:
    discount_input = st.number_input("Diskon (Rp) - mengurangi harga final", value=0, step=1000)

process_btn = st.button("Proses")

if process_btn:
    if not mass_files or pl_file is None or addon_file is None:
        st.warning("Upload Mass Update (minimal 1), Pricelist, dan Addon Mapping dulu ya.")
        st.stop()

    # Read pricelist + addon
    df_pl = read_pricelist_smart(pl_file, sheet_name=0)
    df_add = pd.read_excel(addon_file, sheet_name=0)

    col_sku = find_col_contains(df_pl, ["kodebarang"]) or find_col_contains(df_pl, ["kode barang"])
    col_tot = find_col_contains(df_pl, ["tot"])
    if not col_sku:
        st.error("❌ Pricelist: kolom KODEBARANG tidak ditemukan.")
        st.stop()
    if not col_tot:
        st.error("❌ Pricelist: kolom TOT tidak ditemukan.")
        st.stop()

    sample_price_col = find_col_contains(df_pl, ["m3"]) or find_col_contains(df_pl, ["m4"])
    if not sample_price_col:
        st.error("❌ Pricelist: kolom M3/M4 tidak ditemukan.")
        st.stop()

    price_scale = detect_price_scale_from_pricelist(df_pl, sample_price_col)
    rules = Rules(price_scale=int(price_scale), discount_rp=int(discount_input))
    addon_map = build_addon_map(df_add, price_scale=rules.price_scale)

    issues_all: List[Dict[str, Any]] = []

    # ============
    # MODE: COMBINE (1 file / split per 10k)
    # ============
    if output_mode.startswith("Gabung"):
        try:
            template_wb, template_ws, template_header_row, template_cols, template_sig = create_template_from_first_file(
                mass_files[0].getvalue()
            )
        except Exception as e:
            st.error(f"Gagal baca template Mass Update pertama: {e}")
            st.stop()

        sku_col_template = template_cols["sku"]
        harga_col_template = template_cols["harga"]
        stok_col_template = template_cols.get("stok")  # boleh None

        # Prepare first part workbook
        dst_wb, dst_ws = build_combined_workbook_from_template(template_ws, template_header_row)
        dst_max_col = template_ws.max_column
        dst_write_row = template_header_row + 1
        rows_in_part = 0
        part_index = 1
        part_bytes: List[bytes] = []

        def finalize_current_part():
            nonlocal_rows = rows_in_part  # just for readability
            out = io.BytesIO()
            dst_wb.save(out)
            part_bytes.append(out.getvalue())
            return nonlocal_rows

        for mf in mass_files:
            mp = detect_marketplace_from_filename(mf.name) or "tiktok"
            if mp == "tiktok":
                col_price = find_col_contains(df_pl, ["m3"])
                price_label = "M3"
            else:
                col_price = find_col_contains(df_pl, ["m4"])
                price_label = "M4"

            if not col_price:
                issues_all.append({"file": mf.name, "row": "", "sku_full": "", "base_sku": "", "reason": f"Pricelist: kolom {price_label} tidak ditemukan"})
                continue

            base_price_map, base_stock_map = build_base_maps(
                df_pl=df_pl,
                col_sku=col_sku,
                col_price=col_price,
                col_stock_tot=col_tot,
                price_scale=rules.price_scale
            )

            try:
                wb = load_workbook(io.BytesIO(mf.getvalue()))
                ws = wb.worksheets[0]

                header_row, cols = find_header_row_and_cols(ws, scan_rows=30)
                if header_row == -1:
                    issues_all.append({"file": mf.name, "row": "", "sku_full": "", "base_sku": "", "reason": "Header Mass Update tidak ketemu (SKU/Harga)"})
                    continue

                sig = header_signature(ws, header_row)
                if sig != template_sig:
                    issues_all.append({"file": mf.name, "row": "", "sku_full": "", "base_sku": "", "reason": "Header/kolom berbeda dari template file pertama (skip file)"})
                    continue

                sel_rows, issues = collect_rows_to_update(
                    src_ws=ws,
                    header_row=header_row,
                    sku_col=cols["sku"],
                    base_price_map=base_price_map,
                    addon_map=addon_map,
                    issues_file=mf.name
                )
                issues_all.extend(issues)

                for r in sel_rows:
                    # split kalau sudah 10.000 baris
                    if rows_in_part >= 10000:
                        finalize_current_part()
                        part_index += 1
                        dst_wb, dst_ws = build_combined_workbook_from_template(template_ws, template_header_row)
                        dst_write_row = template_header_row + 1
                        rows_in_part = 0

                    append_row_copy(ws, dst_ws, r, dst_write_row, dst_max_col)

                    sku_val = ws.cell(row=r, column=cols["sku"]).value
                    base, addons = parse_platform_sku(sku_val)

                    base_price = base_price_map.get(base)
                    addon_total, missing = calc_addon_total(addons, addon_map)
                    if base_price is None or missing:
                        # safety
                        dst_write_row += 1
                        rows_in_part += 1
                        continue

                    final_price = int(base_price + addon_total - rules.discount_rp)
                    if final_price < 0:
                        final_price = 0

                    set_cell_value_safe(dst_ws, dst_write_row, harga_col_template, final_price)

                    if stok_col_template is not None:
                        stv = base_stock_map.get(base)
                        if stv is not None:
                            set_cell_value_safe(dst_ws, dst_write_row, stok_col_template, int(stv))

                    dst_write_row += 1
                    rows_in_part += 1

            except Exception as e:
                issues_all.append({"file": mf.name, "row": "", "sku_full": "", "base_sku": "", "reason": f"Gagal proses file: {e}"})
                continue

        # last part
        finalize_current_part()

        # Output
        if len(part_bytes) == 1:
            st.success("✅ Selesai (gabung 1 file)")
            st.download_button(
                "⬇️ Download hasil gabungan (XLSX)",
                data=part_bytes[0],
                file_name="combined_mass_update.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            if issues_all:
                st.warning(f"Ada issues: {len(issues_all)} baris. (lihat issues_report.xlsx)")
                issues_buf = io.BytesIO()
                with pd.ExcelWriter(issues_buf, engine="openpyxl") as w:
                    pd.DataFrame(issues_all).to_excel(w, index=False, sheet_name="issues")
                st.download_button(
                    "⬇️ Download issues_report.xlsx",
                    data=issues_buf.getvalue(),
                    file_name="issues_report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                for i, b in enumerate(part_bytes, start=1):
                    zf.writestr(f"combined_mass_update_part{i}.xlsx", b)
                if issues_all:
                    issues_bytes = io.BytesIO()
                    with pd.ExcelWriter(issues_bytes, engine="openpyxl") as w:
                        pd.DataFrame(issues_all).to_excel(w, index=False, sheet_name="issues")
                    zf.writestr("issues_report.xlsx", issues_bytes.getvalue())

            st.success("✅ Selesai (dibagi beberapa part karena >10.000 baris)")
            st.download_button(
                "⬇️ Download ZIP (hasil gabungan per part)",
                data=zip_buf.getvalue(),
                file_name="combined_mass_update_parts.zip",
                mime="application/zip"
            )

    # ============
    # MODE: PER FILE ZIP
    # ============
    else:
        zip_buf = io.BytesIO()

        with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for mf in mass_files:
                mp = detect_marketplace_from_filename(mf.name) or "tiktok"
                if mp == "tiktok":
                    col_price = find_col_contains(df_pl, ["m3"])
                    price_label = "M3"
                else:
                    col_price = find_col_contains(df_pl, ["m4"])
                    price_label = "M4"

                if not col_price:
                    issues_all.append({"file": mf.name, "row": "", "sku_full": "", "base_sku": "", "reason": f"Pricelist: kolom {price_label} tidak ditemukan"})
                    continue

                base_price_map, base_stock_map = build_base_maps(
                    df_pl=df_pl,
                    col_sku=col_sku,
                    col_price=col_price,
                    col_stock_tot=col_tot,
                    price_scale=rules.price_scale
                )

                try:
                    wb = load_workbook(io.BytesIO(mf.getvalue()))
                    ws = wb.worksheets[0]

                    header_row, cols = find_header_row_and_cols(ws, scan_rows=30)
                    if header_row == -1:
                        issues_all.append({"file": mf.name, "row": "", "sku_full": "", "base_sku": "", "reason": "Header Mass Update tidak ketemu (SKU/Harga)"})
                        continue

                    sel_rows, issues = collect_rows_to_update(
                        src_ws=ws,
                        header_row=header_row,
                        sku_col=cols["sku"],
                        base_price_map=base_price_map,
                        addon_map=addon_map,
                        issues_file=mf.name
                    )
                    issues_all.extend(issues)

                    for r in sel_rows:
                        sku_val = ws.cell(row=r, column=cols["sku"]).value
                        base, addons = parse_platform_sku(sku_val)
                        base_price = base_price_map.get(base)
                        addon_total, missing = calc_addon_total(addons, addon_map)
                        if base_price is None or missing:
                            continue

                        final_price = int(base_price + addon_total - rules.discount_rp)
                        if final_price < 0:
                            final_price = 0

                        set_cell_value_safe(ws, r, cols["harga"], final_price)

                        if "stok" in cols:
                            stv = base_stock_map.get(base)
                            if stv is not None:
                                set_cell_value_safe(ws, r, cols["stok"], int(stv))

                    out = io.BytesIO()
                    wb.save(out)
                    zf.writestr(mf.name.replace(".xlsx", "_updated.xlsx"), out.getvalue())

                except Exception as e:
                    issues_all.append({"file": mf.name, "row": "", "sku_full": "", "base_sku": "", "reason": f"Gagal proses file: {e}"})
                    continue

            if issues_all:
                issues_bytes = io.BytesIO()
                with pd.ExcelWriter(issues_bytes, engine="openpyxl") as w:
                    pd.DataFrame(issues_all).to_excel(w, index=False, sheet_name="issues")
                zf.writestr("issues_report.xlsx", issues_bytes.getvalue())

        st.success("✅ Selesai")
        st.download_button(
            "⬇️ Download ZIP",
            data=zip_buf.getvalue(),
            file_name="mass_update_results.zip",
            mime="application/zip"
        )
